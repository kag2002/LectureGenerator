import csv
import io
import json
import logging
from difflib import SequenceMatcher

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def fuzzy_ratio(a: str, b: str) -> float:
    """Helper to calculate similarity ratio between two strings."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()

def parse_kahoot_file(contents: bytes, filename: str, db_questions: list) -> list[dict]:
    """
    Parse file kết quả Excel hoặc CSV tải về từ Kahoot.
    Sử dụng fuzzy matching để tìm các câu hỏi tương ứng trong DB.
    """
    parsed_rows = []

    if filename.endswith(".xlsx") and OPENPYXL_AVAILABLE:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = None
            for name in wb.sheetnames:
                if "summary" in name.lower() or "question" in name.lower():
                    sheet = wb[name]
                    break
            if not sheet:
                sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                text_cells = [str(cell) for cell in row if cell is not None]
                if not text_cells:
                    continue
                numbers = [cell for cell in row if isinstance(cell, (int, float))]
                
                best_match = None
                best_ratio = 0.0
                for text_val in text_cells:
                    if len(text_val) < 10:
                        continue
                    for db_q in db_questions:
                        ratio = fuzzy_ratio(text_val, db_q.question_text)
                        if ratio > best_ratio:
                            best_ratio = ratio
                            best_match = db_q

                if best_ratio > 0.75 and best_match:
                    correct_val = 0
                    incorrect_val = 0
                    valid_nums = [int(n) for n in numbers if n >= 0]
                    if len(valid_nums) >= 2:
                        correct_val = valid_nums[0]
                        incorrect_val = valid_nums[1]
                    elif len(valid_nums) == 1:
                        correct_val = valid_nums[0]
                        incorrect_val = 0

                    parsed_rows.append({
                        "question_id": best_match.id,
                        "correct_count": correct_val,
                        "incorrect_count": incorrect_val
                    })
        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")

    if not parsed_rows:
        try:
            try:
                decoded = contents.decode("utf-8")
            except Exception:
                decoded = contents.decode("latin-1")

            csv_reader = csv.reader(io.StringIO(decoded))
            for row in csv_reader:
                if not row:
                    continue
                best_match = None
                best_ratio = 0.0
                for cell in row:
                    if len(cell) < 10:
                        continue
                    for db_q in db_questions:
                        ratio = fuzzy_ratio(cell, db_q.question_text)
                        if ratio > best_ratio:
                            best_ratio = ratio
                            best_match = db_q

                if best_ratio > 0.75 and best_match:
                    nums = []
                    for cell in row:
                        try:
                            clean_num = "".join(c for c in cell if c.isdigit())
                            if clean_num:
                                nums.append(int(clean_num))
                        except Exception:
                            pass
                    
                    correct_val = nums[0] if len(nums) > 0 else 0
                    incorrect_val = nums[1] if len(nums) > 1 else 0

                    parsed_rows.append({
                        "question_id": best_match.id,
                        "correct_count": correct_val,
                        "incorrect_count": incorrect_val
                    })
        except Exception as e:
            raise ValueError(f"Không thể đọc file điểm. Vui lòng nạp file CSV hoặc Excel hợp lệ. Lỗi: {str(e)}")

    return parsed_rows

def distribute_incorrect_choices(options_json: str, correct_answer: str, correct_count: int, incorrect_count: int) -> dict:
    """
    Phân phối các câu trả lời sai cho các phương án sai một cách tương đối đồng đều.
    """
    dist = {"A": 0, "B": 0, "C": 0, "D": 0}
    if not options_json:
        return dist

    try:
        options = json.loads(options_json)
        correct_letter = "A"
        for idx, opt in enumerate(options):
            if opt.strip().lower() == correct_answer.strip().lower():
                correct_letter = chr(65 + idx)
                break
        
        dist[correct_letter] = correct_count
        
        wrong_letters = [chr(65 + i) for i in range(len(options)) if chr(65 + i) != correct_letter]
        if wrong_letters and incorrect_count > 0:
            each_wrong = incorrect_count // len(wrong_letters)
            rem = incorrect_count % len(wrong_letters)
            for wl in wrong_letters:
                dist[wl] = each_wrong
            dist[wrong_letters[0]] += rem
    except Exception as e:
        logger.error(f"Error distributing choices: {e}")
        
    return dist

def generate_kahoot_export(questions: list) -> tuple[bytes, str, str]:
    """
    Xuất danh sách câu hỏi thành bytes của file Excel hoặc CSV tương thích với Kahoot template.
    Trả về (file_data, file_extension, media_type).
    """
    import io
    
    headers = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Time limit (seconds)",
        "Correct answer(s)",
    ]

    rows = []
    for q in questions:
        q_text = q.question_text
        if len(q_text) > 95:
            q_text = q_text[:92] + "..."

        opts = []
        if q.options_json:
            try:
                opts = json.loads(q.options_json)
            except Exception:
                opts = []

        while len(opts) < 4:
            opts.append("")

        opts = [opt[:60] if len(opt) > 60 else opt for opt in opts]

        correct_idx = "1"
        correct_val = q.correct_answer.strip().lower()
        found = False
        for idx, opt in enumerate(opts):
            if opt.strip().lower() == correct_val:
                correct_idx = str(idx + 1)
                found = True
                break
        if not found:
            if correct_val in ["a", "b", "c", "d"]:
                idx = ord(correct_val) - ord("a")
                if idx < len(opts):
                    correct_idx = str(idx + 1)

        time_limit = 30
        rows.append([q_text, opts[0], opts[1], opts[2], opts[3], time_limit, correct_idx])

    if OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kahoot Template"
        ws.append(headers)
        for r in rows:
            ws.append(r)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return (
            out.getvalue(),
            "xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        output.seek(0)
        return (
            output.getvalue().encode("utf-8"),
            "csv",
            "text/csv"
        )
